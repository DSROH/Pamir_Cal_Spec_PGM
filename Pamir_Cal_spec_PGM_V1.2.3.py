# %% [markdown]
# > # SM-S711 R11 EUR Auto Cal Spec 변경 PGM
# 

# %% [markdown]
# <font size ="2">
# 
# === Pamir ===
# 
# - V1
# 
#   MIMO/CA Path 추가
# 
#   - MAIN_PRX / DRX: 기본 MAIN path
#   - 4RX_PRX / DRX: 기본 MIMO path
#   - 6RX : CA / ENDC 조건에서 사용하기 위한 RFIC, LNA 예비 port #1의 MAIN path
#   - 12RX : CA / ENDC 조건에서 사용하기 위한 RFIC, LNA 예비 port #1의 MIMO path
#   - 8RX : CA / ENDC 조건에서 사용하기 위한 RFIC, LNA 예비 port #2의 MAIN path
# 
#   ***
# 
# 
#   Pamir 8RX 까지 지원 가능
# 
#   - 10RX : CA / ENDC 조건에서 사용하기 위한 RFIC, LNA 예비 port #2의 MIMO path
#   - 10RX : CA / ENDC 조건에서 사용하기 위한 RFIC, LNA 예비 port #3의 MAIN path
#   - 16RX : CA / ENDC 조건에서 사용하기 위한 RFIC, LNA 예비 port #3의 MIMO path
# 
#   ***
# 
#   - 2G Cal Spec Only option
#     - Rx_AGCOffset 항목 추가
#     - GMSK_Ref_Power, GMSK_Power_TxL, EPSK_Power_TxL 추가
# 
# - V1.1
#   - RFIC Gain cal spec 적용
#   - MTM Calibration log 읽어서 NR/HSPA 따로 Spec 적용
#   - RX Gainstage 넘버 읽어서 Stage 갯수 설정
#   - BW Power Cal 구현 완료
#   - MTM / Daseul Spec 수정 분리/동시 가능하도록 수정
#   - FBRX LPM 구현 예정
# 
# 
# - V1.2
#   - RFIC Gain Cal index 가변에 따라 Data 분리 후 개별 처리
#   - RFIC Gain Cal index 없을 때 keyError -> df.get()으로 Multi-index key 체크 후 있는 경우에만 평균값 반영 or 이전값 - 5
#   - ET_S-APT enable 시 TX, TX2 KeyError 체크 스펙 반영
#   
# - V1.2.1
#   - TX2 RFIC Gain cal 항목 누락 수정
#   - MTM RX Default data 진행 시 RFIC Cal cal data 누락 수정
#   
# - V1.2.2
#   - Data Export 시 Exported_Data Folder 생성 후 저장
#   - Save Data to Excel 시 Format 변경
# 
# - V1.2.3
#   - MTM RX Default data 반영 시 호주 블루틱 offset option 추가
# 
# </font>
# 

# %%
import glob
import os
import threading
import tkinter as tk
import tkinter.scrolledtext as st

import pandas as pd
import ttkbootstrap as ttkbst
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from ttkbootstrap.constants import *

import Common_function as func
import LSI_get_data as getdata

# %%
font_style = Font(
    name="Calibri",
    size=10,
    bold=False,
    italic=False,
    vertAlign=None,  # 첨자
    underline="none",  # 밑줄
    strike=False,  # 취소선
    color="00000000",  # 블랙, # 00FF0000 Red, # 000000FF Blue
)

# %%
def Common_save_Excel(filename, tab1, tab2):
    # Save Data to Excel
    Tabname = filename.replace("Excel_", "")
    Tabname = f"{os.path.splitext(Tabname)[0]}"
    with pd.ExcelWriter(filename) as writer:
        tab1.to_excel(writer, sheet_name=f"{Tabname}_Mean")
        tab2.to_excel(writer, sheet_name=f"{Tabname}_Data")


def WB_Format(filename):
    wb = load_workbook(filename)
    ws = wb.sheetnames
    for sheet in ws:
        col_max = wb[sheet].max_column
        row_max = wb[sheet].max_row
        for i in range(2, row_max + 1, 1):
            for j in range(3, col_max + 1, 1):
                wb[sheet].cell(row=i, column=j).font = font_style
                wb[sheet].cell(row=i, column=j).alignment = Alignment(horizontal="right")
                wb[sheet].cell(row=i, column=j).number_format = "#,##0.0"
                # wb[sheet].cell(row=i, column=j).number_format = builtin_format_code(2)
    wb.save(filename)

# %%
Win_GUI = ttkbst.Window(title="S23FE(R11) LSI Pamir Auto Cal Spec PGM V1.2.3", themename="cosmo")
Win_GUI.attributes("-topmost", True)
Win_GUI.geometry("1425x600")  # py : 1407x560 ipynb : 1635x670
# Win_GUI.option_add("*Font", "Consolas 10")


def change_theme():
    themename = Win_GUI.getvar("themename")
    Win_GUI.style.theme_use(themename)


themes = Win_GUI.style.theme_names()

Left_frame = ttkbst.Frame(Win_GUI)
Left_frame.place(x=0, y=0, width=640, height=600)

# 리스트 프레임
list_frame = ttkbst.Labelframe(Left_frame, text=" Calibration log ", bootstyle=PRIMARY)
list_frame.place(x=5, y=45, width=630, height=85)

scrollbar = ttkbst.Scrollbar(list_frame, orient="vertical")
scrollbar.place(x=600, y=0, width=23, height=62)

list_file = tk.Listbox(list_frame, height=5, yscrollcommand=scrollbar.set)
list_file.place(x=5, y=0, width=595, height=62)

# %%
Scrolled_txt_frame = ttkbst.Frame(Win_GUI)
Scrolled_txt_frame.place(x=640, y=0, width=780, height=600)

text_area = st.ScrolledText(Scrolled_txt_frame, font=("Consolas", 9))
text_area.place(x=0, y=5, width=780, height=545)

right_run_frame = ttkbst.Frame(Win_GUI)
right_run_frame.place(x=640, y=550, width=780, height=50)

Author = ttkbst.Label(right_run_frame, text="dongsub.roh@samsung.com", anchor="w")
Author.place(x=620, y=5, width=160, height=40)

# p_var = tk.DoubleVar()
# progressbar = ttkbst.CTkProgressBar(right_run_frame, variable=p_var, fg_color="#dbdbdb", corner_radius=5)
# progressbar.place(x=0, y=5, width=770, height=45)

# %%
s = ttkbst.Style()
s.configure("info.TButton", font=("Consolas", 10))
s.configure("info.TCheckbutton", font=("Consolas", 10))

daseul_select = ttkbst.BooleanVar()
mtm_select = ttkbst.BooleanVar()
daseul_select.set(False)
mtm_select.set(False)


# 저장 경로 프레임
path_frame = ttkbst.Labelframe(Left_frame, text=" SPC File Path ", bootstyle=PRIMARY)
path_frame.place(x=5, y=135, width=630, height=120)

spc_chkbox = ttkbst.Checkbutton(path_frame, text="SPC", style="info.TCheckbutton", variable=daseul_select)
spc_chkbox.place(x=10, y=10, width=45, height=30)

path_spc_file = ttkbst.Entry(path_frame)
# spc 파일 경로 사전입력
path_spc_file.insert(0, "D:\\")
path_spc_file.place(x=65, y=10, width=455, height=30)

btn_spc = ttkbst.Button(
    path_frame,
    text="SPC (F2)",
    style="info.TButton",
    command=lambda: [func.browse_spc_path(path_spc_file, daseul_select, text_area)],
)
btn_spc.place(x=530, y=10, width=90, height=30)


mtm_chkbox = ttkbst.Checkbutton(path_frame, text="MTM", style="info.TCheckbutton", variable=mtm_select)
mtm_chkbox.place(x=10, y=55, width=45, height=30)

mtm_folder = ttkbst.Entry(path_frame)
# spc 파일 경로 사전입력
mtm_folder.insert(0, "D:\\")
mtm_folder.place(x=65, y=55, width=455, height=30)

btn_mtm = ttkbst.Button(
    path_frame, text="MTM (F3)", style="info.TButton", command=lambda: [func.browse_mtm_path(mtm_folder, mtm_select, text_area)]
)
btn_mtm.place(x=530, y=55, width=90, height=30)


Win_GUI.bind("<F3>", lambda event: [func.browse_mtm_path(mtm_folder, text_area)])

# %%
# 옵션 선택 frame
radio_Btn_frame = ttkbst.Labelframe(Left_frame, text=" Options ", bootstyle=PRIMARY)
radio_Btn_frame.place(x=5, y=260, width=630, height=100)

Option_var = ttkbst.IntVar()

Select_op = "Daseul"
Option_var.set(2)


def Dclick():
    global Select_op
    Select_op = "Daseul"
    Option_var.set(2)


def Mclick():
    global Select_op
    Select_op = "MTM"
    Option_var.set(3)


btn_Option1 = ttkbst.Radiobutton(radio_Btn_frame, text="Cal Spec 조정", value=1, variable=Option_var)
btn_Option1.place(x=10, y=10, width=100, height=25)

btn_Option2 = ttkbst.Radiobutton(radio_Btn_frame, text="Cal 산포 적용", value=2, variable=Option_var)
btn_Option2.place(x=120, y=10, width=100, height=25)
btn_Option2.invoke()

btn_Option3 = ttkbst.Radiobutton(radio_Btn_frame, text="MTM Default Cal Data", value=3, variable=Option_var)
btn_Option3.place(x=230, y=10, width=140, height=25)

Save_data_var = ttkbst.BooleanVar()
chkbox1 = ttkbst.Checkbutton(radio_Btn_frame, text="Save Data to Excel", variable=Save_data_var)
chkbox1.place(x=495, y=10, width=120, height=25)
Save_data_var.set(False)
# chkbox1.configure(state="!selected")

debug_var = ttkbst.BooleanVar()
chkbox2 = ttkbst.Checkbutton(radio_Btn_frame, text="Debug Option", variable=debug_var)
chkbox2.place(x=495, y=45, width=120, height=25)
debug_var.set(False)
# chkbox2.configure(state="!selected")

# ? ======================================== Bluetick for 호주향 ======================================== 
Bluetick_var = ttkbst.BooleanVar()
Bluetick_chkbx = ttkbst.Checkbutton(radio_Btn_frame, text="Bluetick for AU", variable=Bluetick_var)
Bluetick_chkbx.place(x=10, y=45, width=120, height=25)
Bluetick_var.set(True)
# ! There's an even easier way than select() and deselect()! 
# ! If you properly link a checkbutton to a tkinter int or boolean variable, the checkbutton will automatically check and uncheck
# ! if it's given 1/True or 0/False values, respectively

Blue_label_3g = ttkbst.Label(radio_Btn_frame, text="3G", anchor="w")
Blue_label_3g.place(x=120, y=45, width=20, height=25)
Blue_3g_ch = ttkbst.Entry(radio_Btn_frame, justify="right")
Blue_3g_ch.place(x=140, y=45, width=45, height=25)
Blue_3g_ch.insert(tk.END, "4436")

Blue_3g_offs = ttkbst.Entry(radio_Btn_frame, justify="right")
Blue_3g_offs.place(x=190, y=45, width=25, height=25)
Blue_3g_offs.insert(tk.END, "3")
Blue_label_3goffs = ttkbst.Label(radio_Btn_frame, text="dB", anchor="w")
Blue_label_3goffs.place(x=215, y=45, width=20, height=25)

Blue_label_nr = ttkbst.Label(radio_Btn_frame, text="LTE", anchor="w")
Blue_label_nr.place(x=248, y=45, width=25, height=25)
Blue_nr_ch = ttkbst.Entry(radio_Btn_frame, justify="right")
Blue_nr_ch.place(x=273, y=45, width=45, height=25)
Blue_nr_ch.insert(tk.END, "9410")

Blue_nr_offs = ttkbst.Entry(radio_Btn_frame, justify="right")
Blue_nr_offs.place(x=323, y=45, width=25, height=25)
Blue_nr_offs.insert(tk.END, "2")
Blue_label_nroffs = ttkbst.Label(radio_Btn_frame, text="dB", anchor="w")
Blue_label_nroffs.place(x=348, y=45, width=20, height=25)
# ? ======================================== Bluetick for 호주향 ======================================== 

# Cal log 파일 선택
Btn_frame = ttkbst.Frame(Left_frame)
Btn_frame.place(x=5, y=0, width=630, height=45)

btn_add_file1 = ttkbst.Button(
    Btn_frame, text="Daseul log 추가 (F1)", command=lambda: [func.Common_daseul_log(list_file), Dclick()]
)
btn_add_file1.place(x=0, y=5, width=135, height=35)

btn_add_file2 = ttkbst.Button(Btn_frame, text="MTM log 추가 (F4)", command=lambda: [func.Common_mtm_log(list_file), Mclick()])
btn_add_file2.place(x=505, y=5, width=125, height=35)

# Cal log : log 폴더의 CSV 파일 자동 입력
for filename in glob.glob("C:\\DGS\\LOGS\\*.csv"):
    list_file.insert(tk.END, filename)

scrollbar.config(command=list_file.yview)

Win_GUI.bind("<F1>", lambda event: [func.Common_daseul_log(list_file), Dclick()])
Win_GUI.bind("<F4>", lambda event: [func.Common_mtm_log(list_file), Mclick()])

# %%
# Cal Spec frame
cal_spec_frame = ttkbst.Labelframe(Left_frame, text="Cal Spec", bootstyle=PRIMARY)
cal_spec_frame.place(x=5, y=365, width=630, height=185)

Cable_Spec_label = ttkbst.Label(cal_spec_frame, text="Cable Check", anchor="e")
Cable_Spec_label.place(x=10, y=10, width=75, height=25)
Cable_Spec_var = ttkbst.Entry(cal_spec_frame, justify="right")
Cable_Spec_var.place(x=90, y=10, width=40, height=25)
Cable_Spec_var.insert(tk.END, "3")

RFIC_Spec_label = ttkbst.Label(cal_spec_frame, text="RFIC_Gain", anchor="e")
RFIC_Spec_label.place(x=10, y=45, width=75, height=25)
RFIC_Spec_var = ttkbst.Entry(cal_spec_frame, justify="right")
RFIC_Spec_var.place(x=90, y=45, width=40, height=25)
RFIC_Spec_var.insert(tk.END, "5")

APT_Spec_label = ttkbst.Label(cal_spec_frame, text="APT_Spec", anchor="e")
APT_Spec_label.place(x=10, y=80, width=75, height=25)
APT_Spec_var = ttkbst.Entry(cal_spec_frame, justify="right")
APT_Spec_var.place(x=90, y=80, width=40, height=25)
APT_Spec_var.insert(tk.END, "0.5")

BW_Cal_Spec_label = ttkbst.Label(cal_spec_frame, text="BW_Cal", anchor="e")
BW_Cal_Spec_label.place(x=10, y=115, width=75, height=25)
BW_Cal_Spec_var = ttkbst.Entry(cal_spec_frame, justify="right")
BW_Cal_Spec_var.place(x=90, y=115, width=40, height=25)
BW_Cal_Spec_var.insert(tk.END, "3")

FBRX_Spec_label = ttkbst.Label(cal_spec_frame, text="NR FBRX", anchor="e")
FBRX_Spec_label.place(x=140, y=10, width=70, height=25)
FBRX_Spec_var = ttkbst.Entry(cal_spec_frame, justify="right")
FBRX_Spec_var.place(x=215, y=10, width=40, height=25)
FBRX_Spec_var.insert(tk.END, "5")

GMSK_Spec_label = ttkbst.Label(cal_spec_frame, text="GMSK Spec", anchor="e")
GMSK_Spec_label.place(x=140, y=45, width=70, height=25)
GMSK_Spec_var = ttkbst.Entry(cal_spec_frame, justify="right")
GMSK_Spec_var.place(x=215, y=45, width=40, height=25)
GMSK_Spec_var.insert(tk.END, "5")

GTxL_Spec_label = ttkbst.Label(cal_spec_frame, text="GMSK TxL", anchor="e")
GTxL_Spec_label.place(x=140, y=80, width=70, height=25)
GTxL_Spec_var = ttkbst.Entry(cal_spec_frame, justify="right")
GTxL_Spec_var.place(x=215, y=80, width=40, height=25)
GTxL_Spec_var.insert(tk.END, "2")

GCode_Spec_label = ttkbst.Label(cal_spec_frame, text="GMSK Code", anchor="e")
GCode_Spec_label.place(x=140, y=115, width=70, height=25)
GCode_Spec_var = ttkbst.Entry(cal_spec_frame, justify="right")
GCode_Spec_var.place(x=215, y=115, width=40, height=25)
GCode_Spec_var.insert(tk.END, "50")

FBRX_3G_Spec_label = ttkbst.Label(cal_spec_frame, text="3G FBRX", anchor="e")
FBRX_3G_Spec_label.place(x=265, y=10, width=70, height=25)
FBRX_3G_Spec_var = ttkbst.Entry(cal_spec_frame, justify="right")
FBRX_3G_Spec_var.place(x=340, y=10, width=40, height=25)
FBRX_3G_Spec_var.insert(tk.END, "3")

EPSK_Spec_label = ttkbst.Label(cal_spec_frame, text="EPSK Spec", anchor="e")
EPSK_Spec_label.place(x=265, y=45, width=70, height=25)
EPSK_Spec_var = ttkbst.Entry(cal_spec_frame, justify="right")
EPSK_Spec_var.place(x=340, y=45, width=40, height=25)
EPSK_Spec_var.insert(tk.END, "5")

ETxL_Spec_label = ttkbst.Label(cal_spec_frame, text="EPSK TxL", anchor="e")
ETxL_Spec_label.place(x=265, y=80, width=70, height=25)
ETxL_Spec_var = ttkbst.Entry(cal_spec_frame, justify="right")
ETxL_Spec_var.place(x=340, y=80, width=40, height=25)
ETxL_Spec_var.insert(tk.END, "2")

ECode_Spec_label = ttkbst.Label(cal_spec_frame, text="EPSK Code", anchor="e")
ECode_Spec_label.place(x=265, y=115, width=70, height=25)
ECode_Spec_var = ttkbst.Entry(cal_spec_frame, justify="right")
ECode_Spec_var.place(x=340, y=115, width=40, height=25)
ECode_Spec_var.insert(tk.END, "50")

RX_Gain_label = ttkbst.Label(cal_spec_frame, text="NR RX Gain", anchor="e")
RX_Gain_label.place(x=390, y=10, width=70, height=25)
RX_Gain_Spec_var = ttkbst.Entry(cal_spec_frame, justify="right")
RX_Gain_Spec_var.place(x=465, y=10, width=40, height=25)
RX_Gain_Spec_var.insert(tk.END, "5")

RX_Gain_3G_label = ttkbst.Label(cal_spec_frame, text="3G RX Gain", anchor="e")
RX_Gain_3G_label.place(x=390, y=45, width=70, height=25)
RX_Gain_3G_Spec_var = ttkbst.Entry(cal_spec_frame, justify="right")
RX_Gain_3G_Spec_var.place(x=465, y=45, width=40, height=25)
RX_Gain_3G_Spec_var.insert(tk.END, "5")

RX_Gain_2G_label = ttkbst.Label(cal_spec_frame, text="2G RX Gain", anchor="e")
RX_Gain_2G_label.place(x=390, y=80, width=70, height=25)
RX_Gain_2G_Spec_var = ttkbst.Entry(cal_spec_frame, justify="right")
RX_Gain_2G_Spec_var.place(x=465, y=80, width=40, height=25)
RX_Gain_2G_Spec_var.insert(tk.END, "5")

ET_Psat_label = ttkbst.Label(cal_spec_frame, text="ET_Psat", anchor="e")
ET_Psat_label.place(x=515, y=10, width=60, height=25)
ET_Psat_var = ttkbst.Entry(cal_spec_frame, justify="right")
ET_Psat_var.place(x=580, y=10, width=40, height=25)
ET_Psat_var.insert(tk.END, "3")

ET_Pgain_label = ttkbst.Label(cal_spec_frame, text="ET_Pgain", anchor="e")
ET_Pgain_label.place(x=515, y=45, width=60, height=25)
ET_Pgain_var = ttkbst.Entry(cal_spec_frame, justify="right")
ET_Pgain_var.place(x=580, y=45, width=40, height=25)
ET_Pgain_var.insert(tk.END, "3")

ET_Freq_label = ttkbst.Label(cal_spec_frame, text="ET_Freq", anchor="e")
ET_Freq_label.place(x=515, y=80, width=60, height=25)
ET_Freq_var = ttkbst.Entry(cal_spec_frame, justify="right")
ET_Freq_var.place(x=580, y=80, width=40, height=25)
ET_Freq_var.insert(tk.END, "3")

ET_Power_label = ttkbst.Label(cal_spec_frame, text="ET_Power", anchor="e")
ET_Power_label.place(x=515, y=115, width=60, height=25)
ET_Power_var = ttkbst.Entry(cal_spec_frame, justify="right")
ET_Power_var.place(x=580, y=115, width=40, height=25)
ET_Power_var.insert(tk.END, "3")

# %%
# 실행 프레임
left_run_frame = ttkbst.Frame(Left_frame)
left_run_frame.place(x=5, y=555, width=630, height=40)

theme_options = tk.Menubutton(left_run_frame, text="Select a theme")
menu = tk.Menu(theme_options)

for t in themes:
    menu.add_radiobutton(label=t, variable="themename", command=change_theme)

theme_options["menu"] = menu
theme_options.place(x=0, y=5, width=100, height=30)

btn_start = ttkbst.Button(
    left_run_frame,
    text="시작 (F5)",
    command=lambda: [
        threading.Thread(
            target=getdata.start,
            args=(
                list_file,
                Option_var,
                path_spc_file,
                mtm_folder,
                Select_op,
                debug_var,
                daseul_select,
                mtm_select,
                Cable_Spec_var,
                RX_Gain_Spec_var,
                FBRX_Spec_var,
                APT_Spec_var,
                ET_Psat_var,
                ET_Pgain_var,
                ET_Freq_var,
                ET_Power_var,
                BW_Cal_Spec_var,
                RFIC_Spec_var,
                RX_Gain_3G_Spec_var,
                FBRX_3G_Spec_var,
                RX_Gain_2G_Spec_var,
                GMSK_Spec_var,
                GTxL_Spec_var,
                GCode_Spec_var,
                EPSK_Spec_var,
                ETxL_Spec_var,
                ECode_Spec_var,
                Save_data_var,
                Bluetick_var,
                Blue_3g_ch,
                Blue_3g_offs,
                Blue_nr_ch,
                Blue_nr_offs,
                text_area,
            ),
        ).start()
    ],
)
btn_start.place(x=430, y=0, width=200, height=40)

Win_GUI.bind(
    "<F5>",
    lambda event: [
        threading.Thread(
            target=getdata.start,
            args=(
                list_file,
                Option_var,
                path_spc_file,
                mtm_folder,
                Select_op,
                debug_var,
                daseul_select,
                mtm_select,
                Cable_Spec_var,
                RX_Gain_Spec_var,
                FBRX_Spec_var,
                APT_Spec_var,
                ET_Psat_var,
                ET_Pgain_var,
                ET_Freq_var,
                ET_Power_var,
                BW_Cal_Spec_var,
                RFIC_Spec_var,
                RX_Gain_3G_Spec_var,
                FBRX_3G_Spec_var,
                RX_Gain_2G_Spec_var,
                GMSK_Spec_var,
                GTxL_Spec_var,
                GCode_Spec_var,
                EPSK_Spec_var,
                ETxL_Spec_var,
                ECode_Spec_var,
                Save_data_var,
                Bluetick_var,
                Blue_3g_ch,
                Blue_3g_offs,
                Blue_nr_ch,
                Blue_nr_offs,
                text_area,
            ),
        ).start()
    ],
)

Win_GUI.resizable(False, False)
Win_GUI.mainloop()